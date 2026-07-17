"""Utilidades de exportación — genera Excel (xlsx) y PDF desde listas de dicts."""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi.responses import StreamingResponse

# Identidad institucional para los reportes.
BRAND_NAME = "Universidad de Sonsonate"
BRAND_COLOR = "#1E40AF"
LOGO_FILE = Path(__file__).resolve().parents[3] / "static" / "assets" / "uso_logo.png"


_FORMULA_CHARS = ("=", "+", "-", "@", "|", "%")

def _safe_cell(v: Any) -> str:
    """Prefija con apóstrofe valores que Excel interpretaría como fórmula."""
    s = str(v) if v is not None else ""
    if s and s[0] in _FORMULA_CHARS:
        return "'" + s
    return s


def _cell_value(v: Any) -> Any:
    """Números como números para que Excel pueda graficar; el resto se sanea."""
    if isinstance(v, bool):
        return _safe_cell(v)
    if isinstance(v, (int, float)):
        return v
    return _safe_cell(v)


def _num(v: Any) -> float | None:
    """Convierte valores de celda como "1,204" u "87%" a float para graficar."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(str(v).replace("%", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None


# Paleta institucional para gráficas. El primero es el azul de marca.
_CHART_PALETTE = [
    "#2563EB", "#0EA5E9", "#10B981", "#F59E0B",
    "#EF4444", "#8B5CF6", "#14B8A6", "#F97316",
]

def build_excel(
    rows: list[dict[str, Any]],
    sheet_name: str = "Datos",
    title: str | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    if not rows:
        ws.append([BRAND_NAME])
        ws.append(["Sin datos para mostrar."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())
    n_cols = len(headers)

    # ── Membrete: institución, título y metadatos ──
    ws.cell(row=1, column=1, value=BRAND_NAME).font = Font(bold=True, size=13, color="1E40AF")
    ws.cell(row=2, column=1, value=title or sheet_name).font = Font(bold=True, size=11, color="111827")
    ws.cell(row=3, column=1, value=f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}").font = Font(size=9, color="6B7280")
    ws.cell(row=4, column=1, value=f"Total de registros: {len(rows)}").font = Font(size=9, color="6B7280")
    if n_cols > 1:
        for r in range(1, 5):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)

    header_row = 6
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows):
        excel_row = header_row + 1 + r_idx
        fill = "FFFFFF" if r_idx % 2 == 0 else "F1F5F9"
        for col_idx, v in enumerate(row.values(), 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=_safe_cell(v))
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(header_row, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_response(
    rows: list[dict[str, Any]], filename: str, sheet_name: str = "Datos",
    title: str | None = None,
) -> StreamingResponse:
    data = build_excel(rows, sheet_name, title=title)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )



def _draw_header_footer(canvas, doc, *, title: str) -> None:
    """Dibuja el membrete institucional (logo + nombre) arriba y el pie con
    numeración de páginas abajo. Se invoca en cada página del PDF."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    canvas.saveState()
    page_w, page_h = doc.pagesize
    brand = colors.HexColor(BRAND_COLOR)

    # ── Encabezado: banda con logo + institución ──
    top = page_h - 1.2 * cm
    if LOGO_FILE.exists():
        try:
            canvas.drawImage(
                str(LOGO_FILE), 1.5 * cm, top - 0.55 * cm,
                width=1.4 * cm, height=1.4 * cm,
                preserveAspectRatio=True, mask="auto",
            )
            text_x = 1.5 * cm + 1.7 * cm
        except Exception:
            text_x = 1.5 * cm
    else:
        text_x = 1.5 * cm

    canvas.setFillColor(brand)
    canvas.setFont("Helvetica-Bold", 12)
    canvas.drawString(text_x, top + 0.2 * cm, BRAND_NAME)

    # Línea divisoria bajo el membrete
    canvas.setStrokeColor(brand)
    canvas.setLineWidth(1)
    canvas.line(1.5 * cm, top - 0.7 * cm, page_w - 1.5 * cm, top - 0.7 * cm)

    # ── Pie: título a la izquierda, paginación a la derecha ──
    canvas.setFillColor(colors.HexColor("#9CA3AF"))
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(1.5 * cm, 1.1 * cm, title)
    canvas.drawRightString(
        page_w - 1.5 * cm, 1.1 * cm, f"Página {canvas.getPageNumber()}"
    )
    canvas.setStrokeColor(colors.HexColor("#E5E7EB"))
    canvas.setLineWidth(0.5)
    canvas.line(1.5 * cm, 1.4 * cm, page_w - 1.5 * cm, 1.4 * cm)
    canvas.restoreState()


def _table_style():
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_COLOR)),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 8.5),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("LINEBELOW",  (0, 0), (-1, 0), 0.5, colors.HexColor(BRAND_COLOR)),
        ("LINEBELOW",  (0, 1), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])


def build_pdf(
    rows: list[dict[str, Any]],
    title: str = "Reporte",
    subtitle: str | None = None,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    buf = io.BytesIO()
    page = landscape(A4) if rows and len(rows[0]) > 5 else A4
    # Margen superior amplio para dejar espacio al membrete dibujado.
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=2.6 * cm, bottomMargin=1.8 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=15,
        textColor=colors.HexColor("#111827"), spaceAfter=2,
    )
    meta_style = ParagraphStyle(
        "ReportMeta", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#6B7280"), spaceAfter=1,
    )

    story: list = [Paragraph(title, title_style)]
    if subtitle:
        story.append(Paragraph(subtitle, meta_style))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", meta_style
    ))
    story.append(Paragraph(f"Total de registros: {len(rows)}", meta_style))
    story.append(Spacer(1, 0.5 * cm))

    first_nonempty = next((r for r in rows if r), None)
    if not first_nonempty:
        story.append(Paragraph("Sin datos para mostrar.", meta_style))
    else:
        headers = list(first_nonempty.keys())
        col_count = len(headers)
        col_width = (page[0] - 3 * cm) / col_count
        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in rows]
        table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        table.setStyle(_table_style())
        story.append(table)

    def _decorate(canvas, d):
        _draw_header_footer(canvas, d, title=title)

    doc.build(story, onFirstPage=_decorate, onLaterPages=_decorate)
    return buf.getvalue()


def pdf_response(
    rows: list[dict[str, Any]],
    filename: str,
    title: str = "Reporte",
    subtitle: str | None = None,
) -> StreamingResponse:
    data = build_pdf(rows, title, subtitle)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )


def _chart_series(rows: list[dict], spec: dict) -> tuple[list[str], list[float]] | None:
    """Extrae (etiquetas, valores) según el spec de gráfica.

    Dos modos:
      {"label": "Tema", "value": "Consultas"}          → una fila por punto
      {"columns": ["Positivas", "Negativas"]}           → columnas de la primera fila
    """
    if spec.get("columns"):
        if not rows:
            return None
        first = rows[0]
        labels, values = [], []
        for col in spec["columns"]:
            n = _num(first.get(col))
            if n is not None:
                labels.append(str(col))
                values.append(n)
        return (labels, values) if values else None

    label_key, value_key = spec.get("label"), spec.get("value")
    if not label_key or not value_key:
        return None
    labels, values = [], []
    for row in rows:
        n = _num(row.get(value_key))
        if n is None:
            continue
        labels.append(str(row.get(label_key, "")))
        values.append(n)
    return (labels, values) if values else None


def _chart_drawing(rows: list[dict], spec: dict, avail_width: float):
    """Construye una gráfica (Drawing de reportlab) desde las filas de la sección.

    spec: {"type": "line"|"bar"|"pie", ...claves de _chart_series}
    Devuelve None si no hay datos numéricos suficientes.
    """
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    series = _chart_series(rows, spec)
    if not series or len(series[1]) < 2 or sum(series[1]) <= 0:
        return None
    labels, values = series
    kind = spec.get("type", "bar")
    brand = colors.HexColor(BRAND_COLOR)

    if kind == "pie":
        height = 4.6 * cm
        d = Drawing(avail_width, height)
        pie = Pie()
        pie.x = avail_width / 2 - 1.7 * cm
        pie.y = 0.5 * cm
        pie.width = pie.height = 3.4 * cm
        pie.data = values
        total = sum(values) or 1
        pie.labels = [
            f"{lab}: {val:g} ({val / total * 100:.0f}%)"
            for lab, val in zip(labels, values)
        ]
        pie.sideLabels = True
        pie.slices.fontSize = 7
        pie.slices.strokeColor = colors.white
        pie.slices.strokeWidth = 0.5
        for i in range(len(values)):
            pie.slices[i].fillColor = colors.HexColor(_CHART_PALETTE[i % len(_CHART_PALETTE)])
        d.add(pie)
        return d

    height = 5.2 * cm
    d = Drawing(avail_width, height)
    chart = HorizontalLineChart() if kind == "line" else VerticalBarChart()
    chart.x = 1.2 * cm
    chart.y = 1.1 * cm
    chart.width = avail_width - 1.8 * cm
    chart.height = height - 1.5 * cm
    chart.data = [tuple(values)]
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 6.5
    chart.valueAxis.labels.fillColor = colors.HexColor("#6B7280")
    chart.valueAxis.strokeColor = colors.HexColor("#D1D5DB")

    # Con muchas categorías se muestran solo unas 10 etiquetas.
    step = max(1, len(labels) // 10)
    cat_names = [
        (lab[:18] if i % step == 0 else "") for i, lab in enumerate(labels)
    ] if kind == "line" else [lab[:18] for lab in labels]
    chart.categoryAxis.categoryNames = cat_names
    chart.categoryAxis.labels.fontSize = 6.5
    chart.categoryAxis.labels.fillColor = colors.HexColor("#6B7280")
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.boxAnchor = "ne"
    chart.categoryAxis.labels.dy = -2
    chart.categoryAxis.strokeColor = colors.HexColor("#D1D5DB")

    if kind == "line":
        chart.lines[0].strokeColor = brand
        chart.lines[0].strokeWidth = 1.6
    else:
        chart.bars[0].fillColor = brand
        chart.bars[0].strokeColor = None
        chart.barLabels.fontSize = 6
        chart.barLabels.fillColor = colors.HexColor("#374151")
        chart.barLabelFormat = "%.0f"
        chart.barLabels.nudge = 6

    d.add(String(0, height - 8, "", fontSize=1))
    return d


def _cover_story(title: str, subtitle: str | None, page_size) -> list:
    """Portada institucional: logo, institución, título del reporte y período."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import HRFlowable, Image, PageBreak, Paragraph, Spacer

    center = dict(alignment=TA_CENTER, fontName="Helvetica")
    brand_style = ParagraphStyle("CoverBrand", fontSize=20, leading=24,
                                 textColor=colors.HexColor(BRAND_COLOR),
                                 fontName="Helvetica-Bold", alignment=TA_CENTER)
    sub_style = ParagraphStyle("CoverSub", fontSize=10, leading=14,
                               textColor=colors.HexColor("#6B7280"), **center)
    title_style = ParagraphStyle("CoverTitle", fontSize=26, leading=32,
                                 textColor=colors.HexColor("#111827"),
                                 fontName="Helvetica-Bold", alignment=TA_CENTER)
    meta_style = ParagraphStyle("CoverMeta", fontSize=11, leading=16,
                                textColor=colors.HexColor("#374151"), **center)

    story: list = [Spacer(1, 3.2 * cm)]
    if LOGO_FILE.exists():
        try:
            story.append(Image(str(LOGO_FILE), width=3.6 * cm, height=3.6 * cm, kind="proportional"))
            story.append(Spacer(1, 0.8 * cm))
        except Exception:
            pass
    story.append(Paragraph(BRAND_NAME, brand_style))
    story.append(Spacer(1, 1.6 * cm))
    story.append(HRFlowable(width="40%", thickness=2, color=colors.HexColor(BRAND_COLOR)))
    story.append(Spacer(1, 1.2 * cm))
    story.append(Paragraph(title, title_style))
    if subtitle:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(subtitle, meta_style))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", sub_style
    ))
    story.append(PageBreak())
    return story


def build_pdf_report(
    sections: list[dict[str, Any]],
    title: str = "Reporte",
    subtitle: str | None = None,
) -> bytes:
    """Genera un PDF con portada, secciones de texto, tablas y gráficas.

    Cada sección admite:
      {"title", "rows"}            → tabla
      {"title", "text"}            → párrafo narrativo (p. ej. resumen ejecutivo)
      {"title", "rows", "chart"}   → gráfica sobre las filas + tabla debajo
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    nonempty = [s for s in sections if s.get("rows") or s.get("text")]
    max_cols = max((len(s["rows"][0]) for s in nonempty if s.get("rows")), default=0)
    page_size = landscape(A4) if max_cols > 5 else A4

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=2.6 * cm, bottomMargin=1.8 * cm,
    )

    styles = getSampleStyleSheet()
    sec_style = ParagraphStyle(
        "SecTitle",
        parent=styles["Normal"],
        fontSize=10,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor(BRAND_COLOR),
        spaceBefore=10,
        spaceAfter=3,
    )
    body_style = ParagraphStyle(
        "SecBody", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9.5, leading=15,
        textColor=colors.HexColor("#1F2937"), spaceAfter=4,
    )

    story: list = _cover_story(title, subtitle, page_size)

    page_width = page_size[0] - 3 * cm

    for section in nonempty:
        sec_title = section.get("title", "")
        story.append(Paragraph(sec_title, sec_style))
        story.append(HRFlowable(
            width="100%", thickness=0.5,
            color=colors.HexColor("#2563EB"), spaceAfter=3,
        ))

        if section.get("text"):
            for para in str(section["text"]).split("\n"):
                if para.strip():
                    story.append(Paragraph(para.strip(), body_style))
            story.append(Spacer(1, 0.3 * cm))
            continue

        rows = section["rows"]

        chart_spec = section.get("chart")
        if chart_spec:
            drawing = _chart_drawing(rows, chart_spec, page_width)
            if drawing is not None:
                story.append(Spacer(1, 0.15 * cm))
                story.append(drawing)
                story.append(Spacer(1, 0.2 * cm))

        headers = list(rows[0].keys())
        col_count = len(headers)
        col_width = page_width / col_count

        table_data = [headers] + [
            [str(row.get(h, "") or "") for h in headers] for row in rows
        ]
        table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 8),
            ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",   (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",  (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * cm))

    def _decorate(canvas, d):
        _draw_header_footer(canvas, d, title=title)

    def _decorate_cover(canvas, d):
        from reportlab.lib.units import cm as _cm
        canvas.saveState()
        canvas.setFillColor(colors.HexColor(BRAND_COLOR))
        canvas.rect(0, 0, d.pagesize[0], 0.6 * _cm, stroke=0, fill=1)
        canvas.restoreState()

    doc.build(story, onFirstPage=_decorate_cover, onLaterPages=_decorate)
    return buf.getvalue()


def pdf_report_response(
    sections: list[dict[str, Any]],
    filename: str,
    title: str = "Reporte",
    subtitle: str | None = None,
) -> StreamingResponse:
    data = build_pdf_report(sections, title, subtitle)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )
