"""Tests unitarios directos para app/services/ingestion/export.py.

Se llama a las funciones puras (build_excel, build_pdf, build_pdf_report y
helpers internos de gráficas) directamente, sin pasar por ningún endpoint,
y se verifica el contenido real generado (con openpyxl para xlsx, y
parseo básico de estructura para PDF).

Foco de cobertura pedido:
- export.py líneas 320-384 (_chart_drawing: pie/bar/line, casos sin datos
  suficientes, muchas categorías) y 291-301 (_chart_series con
  {"label", "value"}).
- También se cubren ramas de build_excel/build_pdf/build_pdf_report con
  filas vacías, formulas peligrosas en celdas, y portada con/sin logo.
"""
from __future__ import annotations

import io

import pytest

from app.services.ingestion import export as export_mod
from app.services.ingestion.export import (
    _cell_value,
    _chart_drawing,
    _chart_series,
    _num,
    _safe_cell,
    build_excel,
    build_pdf,
    build_pdf_report,
)


# ---------------------------------------------------------------------------
# _safe_cell / _cell_value / _num
# ---------------------------------------------------------------------------

class TestSafeCell:
    def test_prefixes_formula_chars(self):
        for prefix in ("=", "+", "-", "@", "|", "%"):
            assert _safe_cell(f"{prefix}SUM(A1)") == f"'{prefix}SUM(A1)"

    def test_leaves_normal_string_untouched(self):
        assert _safe_cell("hola mundo") == "hola mundo"

    def test_none_becomes_empty_string(self):
        assert _safe_cell(None) == ""

    def test_empty_string_stays_empty(self):
        assert _safe_cell("") == ""


class TestCellValue:
    def test_bool_is_sanitized_as_string(self):
        assert _cell_value(True) == "True"
        assert _cell_value(False) == "False"

    def test_int_and_float_pass_through_as_numbers(self):
        assert _cell_value(42) == 42
        assert _cell_value(3.14) == 3.14

    def test_string_is_sanitized(self):
        assert _cell_value("=cmd") == "'=cmd"


class TestNum:
    def test_int_and_float_passthrough(self):
        assert _num(5) == 5.0
        assert _num(2.5) == 2.5

    def test_bool_is_not_treated_as_number(self):
        # bool es subclase de int; la función excluye explícitamente bool.
        assert _num(True) is None

    def test_parses_percentage_and_thousands_separator(self):
        assert _num("87%") == 87.0
        assert _num("1,204") == 1204.0
        assert _num(" 42 ") == 42.0

    def test_unparseable_returns_none(self):
        assert _num("no-es-numero") is None
        assert _num(None) is None


# ---------------------------------------------------------------------------
# build_excel
# ---------------------------------------------------------------------------

class TestBuildExcel:
    def test_empty_rows_produces_placeholder_sheet(self):
        data = build_excel([], sheet_name="VacioSheet", title="Reporte vacio")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A1"].value == export_mod.BRAND_NAME
        assert ws["A2"].value == "Sin datos para mostrar."

    def test_rows_written_with_header_and_metadata(self):
        rows = [
            {"Tema": "Matriculas", "Consultas": 10},
            {"Tema": "Becas", "Consultas": 5},
        ]
        data = build_excel(rows, sheet_name="Datos", title="Mi reporte")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws["A1"].value == export_mod.BRAND_NAME
        assert ws["A2"].value == "Mi reporte"
        assert "Total de registros: 2" in ws["A4"].value
        # Encabezados en la fila 6.
        assert ws.cell(row=6, column=1).value == "Tema"
        assert ws.cell(row=6, column=2).value == "Consultas"
        # Filas de datos.
        assert ws.cell(row=7, column=1).value == "Matriculas"
        assert ws.cell(row=7, column=2).value == "10"

    def test_formula_like_value_is_escaped_in_cell(self):
        rows = [{"Campo": "=2+2"}]
        data = build_excel(rows)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=7, column=1).value == "'=2+2"

    def test_sheet_name_truncated_to_31_chars(self):
        long_name = "x" * 50
        data = build_excel([{"a": 1}], sheet_name=long_name)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.active.title) == 31


# ---------------------------------------------------------------------------
# build_pdf
# ---------------------------------------------------------------------------

class TestBuildPdf:
    def test_empty_rows_produces_valid_pdf_bytes(self):
        data = build_pdf([], title="Reporte vacio")
        assert data[:4] == b"%PDF"

    def test_rows_with_few_columns_uses_portrait(self):
        rows = [{"a": 1, "b": 2}]
        data = build_pdf(rows, title="Chico")
        assert data[:4] == b"%PDF"

    def test_rows_with_many_columns_uses_landscape(self):
        rows = [{f"col{i}": i for i in range(8)}]
        data = build_pdf(rows, title="Grande")
        assert data[:4] == b"%PDF"

    def test_subtitle_included_does_not_break_build(self):
        data = build_pdf([{"a": 1}], title="T", subtitle="Sub")
        assert data[:4] == b"%PDF"

    def test_rows_all_empty_dicts_triggers_no_data_branch(self):
        # first_nonempty es None -> rama "Sin datos para mostrar."
        data = build_pdf([{}, {}], title="Todo vacio")
        assert data[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# _chart_series — foco en líneas 291-301 y 279-289
# ---------------------------------------------------------------------------

class TestChartSeries:
    def test_columns_mode_extracts_first_row_values(self):
        rows = [{"Positivas": 10, "Negativas": 5}, {"Positivas": 99, "Negativas": 99}]
        spec = {"columns": ["Positivas", "Negativas"]}
        labels, values = _chart_series(rows, spec)
        assert labels == ["Positivas", "Negativas"]
        assert values == [10.0, 5.0]

    def test_columns_mode_empty_rows_returns_none(self):
        assert _chart_series([], {"columns": ["A"]}) is None

    def test_columns_mode_no_numeric_values_returns_none(self):
        rows = [{"A": "texto"}]
        assert _chart_series(rows, {"columns": ["A"]}) is None

    def test_label_value_mode_extracts_series(self):
        rows = [
            {"Tema": "Matriculas", "Consultas": 10},
            {"Tema": "Becas", "Consultas": "5"},
        ]
        spec = {"label": "Tema", "value": "Consultas"}
        labels, values = _chart_series(rows, spec)
        assert labels == ["Matriculas", "Becas"]
        assert values == [10.0, 5.0]

    def test_label_value_mode_skips_non_numeric_rows(self):
        rows = [
            {"Tema": "A", "Consultas": "no-numero"},
            {"Tema": "B", "Consultas": 3},
        ]
        spec = {"label": "Tema", "value": "Consultas"}
        labels, values = _chart_series(rows, spec)
        assert labels == ["B"]
        assert values == [3.0]

    def test_missing_label_or_value_key_returns_none(self):
        rows = [{"Tema": "A"}]
        assert _chart_series(rows, {"label": "Tema"}) is None
        assert _chart_series(rows, {"value": "Consultas"}) is None
        assert _chart_series(rows, {}) is None

    def test_label_value_mode_all_non_numeric_returns_none(self):
        rows = [{"Tema": "A", "Consultas": "x"}]
        spec = {"label": "Tema", "value": "Consultas"}
        assert _chart_series(rows, spec) is None


# ---------------------------------------------------------------------------
# _chart_drawing — foco principal: líneas 320-384
# ---------------------------------------------------------------------------

class TestChartDrawing:
    def test_insufficient_data_returns_none_single_point(self):
        rows = [{"Tema": "A", "Consultas": 10}]
        spec = {"type": "bar", "label": "Tema", "value": "Consultas"}
        assert _chart_drawing(rows, spec, avail_width=400) is None

    def test_all_zero_values_returns_none(self):
        rows = [{"Tema": "A", "Consultas": 0}, {"Tema": "B", "Consultas": 0}]
        spec = {"type": "bar", "label": "Tema", "value": "Consultas"}
        assert _chart_drawing(rows, spec, avail_width=400) is None

    def test_no_series_returns_none(self):
        spec = {"type": "bar", "label": "Tema", "value": "Consultas"}
        assert _chart_drawing([], spec, avail_width=400) is None

    def test_pie_chart_builds_drawing_with_slices(self):
        rows = [{"Positivas": 70, "Negativas": 30}]
        spec = {"type": "pie", "columns": ["Positivas", "Negativas"]}
        d = _chart_drawing(rows, spec, avail_width=400)
        assert d is not None
        # El Drawing debe contener el objeto Pie agregado.
        from reportlab.graphics.charts.piecharts import Pie
        assert any(isinstance(c, Pie) for c in d.contents)

    def test_bar_chart_is_added_to_drawing(self):
        # Regresión del bug donde VerticalBarChart se configuraba por completo
        # (datos, ejes, barras, etiquetas) pero nunca se agregaba al Drawing
        # via d.add(chart) — el reporte exportado se veía en blanco. Corregido
        # en export.py: ahora sí se agrega, igual que la rama pie (d.add(pie)).
        rows = [
            {"Tema": "Matriculas", "Consultas": 10},
            {"Tema": "Becas", "Consultas": 20},
            {"Tema": "Horarios", "Consultas": 5},
        ]
        spec = {"type": "bar", "label": "Tema", "value": "Consultas"}
        d = _chart_drawing(rows, spec, avail_width=400)
        assert d is not None
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        assert any(isinstance(c, VerticalBarChart) for c in d.contents)

    def test_line_chart_is_added_to_drawing(self):
        # Misma regresión que el bar chart (ver test anterior).
        rows = [{"Dia": f"D{i}", "Valor": i} for i in range(1, 6)]
        spec = {"type": "line", "label": "Dia", "value": "Valor"}
        d = _chart_drawing(rows, spec, avail_width=400)
        assert d is not None
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        assert any(isinstance(c, HorizontalLineChart) for c in d.contents)

    def test_line_chart_with_many_categories_thins_labels(self):
        # step = len(labels)//10 > 1 -> ejercita la rama de "" en cat_names.
        rows = [{"Dia": f"D{i}", "Valor": i + 1} for i in range(25)]
        spec = {"type": "line", "label": "Dia", "value": "Valor"}
        d = _chart_drawing(rows, spec, avail_width=400)
        assert d is not None

    def test_default_type_is_bar_when_type_missing(self):
        rows = [{"Tema": "A", "Consultas": 1}, {"Tema": "B", "Consultas": 2}]
        spec = {"label": "Tema", "value": "Consultas"}
        d = _chart_drawing(rows, spec, avail_width=400)
        assert d is not None
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        assert any(isinstance(c, VerticalBarChart) for c in d.contents)


# ---------------------------------------------------------------------------
# build_pdf_report — foco 291-301/320-384 integrados + 412-413, 502-504
# ---------------------------------------------------------------------------

class TestBuildPdfReport:
    def test_report_with_table_section_only(self):
        sections = [
            {"title": "Consultas por tema", "rows": [
                {"Tema": "Matriculas", "Consultas": 10},
                {"Tema": "Becas", "Consultas": 5},
            ]},
        ]
        data = build_pdf_report(sections, title="Reporte de uso")
        assert data[:4] == b"%PDF"

    def test_report_with_text_section(self):
        sections = [
            {"title": "Resumen ejecutivo", "text": "Primera linea.\nSegunda linea."},
        ]
        data = build_pdf_report(sections, title="Resumen")
        assert data[:4] == b"%PDF"

    def test_report_with_chart_section_bar(self):
        sections = [
            {
                "title": "Temas mas consultados",
                "rows": [
                    {"Tema": "Matriculas", "Consultas": 10},
                    {"Tema": "Becas", "Consultas": 20},
                ],
                "chart": {"type": "bar", "label": "Tema", "value": "Consultas"},
            },
        ]
        data = build_pdf_report(sections, title="Con grafica")
        assert data[:4] == b"%PDF"

    def test_report_with_chart_section_that_yields_no_drawing(self):
        # Solo 1 fila -> _chart_drawing devuelve None -> no debe fallar el build,
        # solo omite la gráfica (rama `if drawing is not None` en 501-504).
        sections = [
            {
                "title": "Insuficiente",
                "rows": [{"Tema": "Unico", "Consultas": 1}],
                "chart": {"type": "bar", "label": "Tema", "value": "Consultas"},
            },
        ]
        data = build_pdf_report(sections, title="Sin grafica valida")
        assert data[:4] == b"%PDF"

    def test_report_skips_fully_empty_sections(self):
        sections = [
            {"title": "Vacia", "rows": []},
            {"title": "Con datos", "rows": [{"a": 1, "b": 2}]},
        ]
        data = build_pdf_report(sections, title="Mixto")
        assert data[:4] == b"%PDF"

    def test_report_uses_landscape_when_many_columns(self):
        sections = [
            {"title": "Ancha", "rows": [{f"c{i}": i for i in range(8)}]},
        ]
        data = build_pdf_report(sections, title="Ancho")
        assert data[:4] == b"%PDF"

    def test_report_multiple_sections_mixed(self):
        sections = [
            {"title": "Resumen", "text": "Texto narrativo."},
            {
                "title": "Detalle",
                "rows": [{"Tema": "A", "N": 3}, {"Tema": "B", "N": 7}],
                "chart": {"type": "pie", "columns": ["N"]},
            },
        ]
        data = build_pdf_report(sections, title="Completo", subtitle="Periodo julio")
        assert data[:4] == b"%PDF"

    def test_no_logo_file_skips_image_in_cover(self, monkeypatch):
        from pathlib import Path
        monkeypatch.setattr(export_mod, "LOGO_FILE", Path("/no/existe/logo.png"))
        sections = [{"title": "T", "rows": [{"a": 1}]}]
        data = build_pdf_report(sections, title="Sin logo")
        assert data[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Response wrappers (streaming) — sanity de integración liviana
# ---------------------------------------------------------------------------

class TestResponseWrappers:
    def test_excel_response_has_correct_media_type_and_filename(self):
        from app.services.ingestion.export import excel_response
        resp = excel_response([{"a": 1}], filename="mi_reporte")
        assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert 'mi_reporte.xlsx' in resp.headers["content-disposition"]

    def test_pdf_response_has_correct_media_type_and_filename(self):
        from app.services.ingestion.export import pdf_response
        resp = pdf_response([{"a": 1}], filename="mi_reporte")
        assert resp.media_type == "application/pdf"
        assert 'mi_reporte.pdf' in resp.headers["content-disposition"]

    def test_pdf_report_response_has_correct_media_type_and_filename(self):
        from app.services.ingestion.export import pdf_report_response
        resp = pdf_report_response(
            [{"title": "T", "rows": [{"a": 1}]}], filename="reporte_completo",
        )
        assert resp.media_type == "application/pdf"
        assert 'reporte_completo.pdf' in resp.headers["content-disposition"]
